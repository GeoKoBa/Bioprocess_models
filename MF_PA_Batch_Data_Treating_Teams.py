import pandas as pd
import numpy as np
from scipy.signal import savgol_filter, find_peaks
import matplotlib.pyplot as plt
from datetime import datetime
import os
import glob
from scipy.optimize import curve_fit
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, NamedStyle
import re
import logging
import pickle

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def fit_phase_equation(x, y, phase_name):
    """Fit appropriate equation based on phase type"""
    try:
        if phase_name == "Lag Phase":
            # Exponential growth starting from low values
            def lag_func(x, a, b, c):
                return a * np.exp(b * x) + c
            popt, _ = curve_fit(lag_func, x, y, maxfev=2000)
            return "y = {:.2f} * exp({:.2f}x) + {:.2f}".format(*popt), lag_func, popt
            
        elif phase_name == "Exponential Phase":
            # Exponential growth
            def exp_func(x, a, b):
                return a * np.exp(b * x)
            popt, _ = curve_fit(exp_func, x, y, maxfev=2000)
            return "y = {:.2f} * exp({:.2f}x)".format(*popt), exp_func, popt
            
        elif phase_name == "Production Phase":
            # Polynomial fit
            coeffs = np.polyfit(x, y, 2)
            def prod_func(x):
                return coeffs[0] * x**2 + coeffs[1] * x + coeffs[2]
            return "y = {:.2f}x² + {:.2f}x + {:.2f}".format(*coeffs), prod_func, coeffs
            
        elif phase_name == "Death Phase":
            # Exponential decay
            def death_func(x, a, b, c):
                return a * np.exp(-b * x) + c
            popt, _ = curve_fit(death_func, x, y, maxfev=2000)
            return "y = {:.2f} * exp(-{:.2f}x) + {:.2f}".format(*popt), death_func, popt
            
        elif phase_name == "Post Death Phase":
            # Linear asymptotic
            def post_death_func(x, a, b):
                return a / (1 + np.exp(-b * x))
            popt, _ = curve_fit(post_death_func, x, y, maxfev=2000)
            return "y = {:.2f} / (1 + exp(-{:.2f}x))".format(*popt), post_death_func, popt
            
    except Exception as e:
        logging.error(f"Error fitting {phase_name}: {str(e)}")
        return None, None, None
    
    return None, None, None

def analyze_fermentation_phases(csv_file, value_column='mass flow (CO2)', 
                              timestamp_column='Timestamp', smoothing_window=50, 
                              peak_prominence=1.0):
    """
    Analyze fermentation phases from mass flow data using curve characteristics.
    """
    # Load and process mass flow data
    df = pd.read_csv(csv_file, delimiter=';')
    df[timestamp_column] = pd.to_datetime(df[timestamp_column]).dt.tz_localize(None)
    
    # Convert timestamps to hours from start
    start_time = df[timestamp_column].min()
    df['hours'] = (df[timestamp_column] - start_time).dt.total_seconds() / 3600
    
    # Set hours as index and sort
    df = df.set_index('hours')
    df = df.sort_index()
    df = df.ffill()

    # Smooth the data
    df['smoothed'] = df[value_column].rolling(window=smoothing_window, center=True).mean()
    df['smoothed'] = df['smoothed'].ffill().bfill()

    # Calculate derivatives
    df['first_derivative'] = np.gradient(df['smoothed'].values)
    df['rolling_derivative'] = df['first_derivative'].rolling(window=smoothing_window, center=True).mean()
    df['second_derivative'] = np.gradient(df['rolling_derivative'].values)
    
    # Find peaks
    peaks, properties = find_peaks(df['smoothed'].values, prominence=peak_prominence)
    
    if len(peaks) == 0:
        main_peak_idx = len(df) // 2
    else:
        main_peak_idx = peaks[np.argmax(properties['prominences'])]
    
    def find_derivative_discontinuity(data, start_idx, window_size=20):
        """Find significant discontinuity in derivative after the peak"""
        derivatives = np.abs(data['first_derivative'].iloc[start_idx:])
        rolling_std = derivatives.rolling(window=window_size).std()
        
        spike_threshold = np.percentile(rolling_std, 95)
        potential_spikes = np.where(rolling_std > spike_threshold)[0]
        valid_spikes = [idx for idx in potential_spikes if data.index[start_idx + idx] > 25]
        
        # Check new constraints separately for production phase end
        for idx in range(start_idx, len(data)):
            if data['mass flow (CO2)'].iloc[idx] < 3 and np.abs(data['rolling_derivative'].iloc[idx]) < 0.00075:
                return idx

        # Retain existing logic for slope-based phase transitions
        if len(valid_spikes) > 0:
            return start_idx + valid_spikes[0]
        
        slope_changes = np.abs(data['second_derivative'].iloc[start_idx:])
        significant_changes = np.where(slope_changes > np.percentile(slope_changes, 90))[0]
        valid_changes = [idx for idx in significant_changes if data.index[start_idx + idx] > 250]

        if len(valid_changes) > 0:
            return start_idx + valid_changes[0]
        
        return None

    def find_post_death_transition(data, start_idx, asymptote_threshold=0.4, window_size=20):
        """
        Find transition to post-death phase when mass flow becomes asymptotic to zero.
        """
        for i in range(start_idx, len(data) - window_size):
            window_values = data['smoothed'].iloc[i:i+window_size]
            window_derivatives = np.abs(data['rolling_derivative'].iloc[i:i+window_size])
            
            if (np.all(window_values < asymptote_threshold) and 
                np.mean(window_derivatives) < np.std(data['smoothed']) * 0.1):
                return i
        
        return len(data) - 1
    
    # Phase detection parameters
    lag_threshold = np.std(df['rolling_derivative']) * 3
    
    # Find phase transition points
    lag_end = 0
    for i in range(len(df)):
        if df['rolling_derivative'].iloc[i] > lag_threshold:
            lag_end = i
            break
    
    exp_end = main_peak_idx
    prod_end = find_derivative_discontinuity(df, main_peak_idx)
    if prod_end is None:
        prod_end = len(df) - 1  # Default to the end of the data if no discontinuity is found
    post_death_start = find_post_death_transition(df, prod_end)
    
    # Create phase information with trendline equations
    phase_info = []
    phases = [
        ('Lag Phase', 0, lag_end),
        ('Exponential Phase', lag_end, exp_end),
        ('Production Phase', exp_end, prod_end),
        ('Death Phase', prod_end, post_death_start),
        ('Post Death Phase', post_death_start, len(df)-1)
    ]
    
    for phase_name, start_idx, end_idx in phases:
        phase_data = df.iloc[start_idx:end_idx+1]
        mean_rate = np.mean(phase_data['rolling_derivative'])
        peak_rate = np.max(np.abs(phase_data['rolling_derivative']))
        
        # Fit trendline for this phase
        x = phase_data.index.values - phase_data.index.values[0]  # Normalize x to start at 0
        y = phase_data['smoothed'].values
        equation, func, params = fit_phase_equation(x, y, phase_name)
        
        phase_info.append({
            'Batch': os.path.basename(csv_file),
            'Phase': phase_name,
            'Start Hour': df.index[start_idx],
            'End Hour': df.index[end_idx],
            'Duration (hours)': df.index[end_idx] - df.index[start_idx],
            'Start Value': df[value_column].iloc[start_idx],
            'End Value': df[value_column].iloc[end_idx],
            'Mean Rate': mean_rate,
            'Peak Rate': peak_rate,
            'Trendline Equation': equation if equation else "Could not fit equation"
        })
    
    return phase_info, df

def analyze_single_batch(csv_file):
    """Analyze a single CSV file and create a report"""
    if not os.path.exists(csv_file):
        print(f"The file {csv_file} does not exist.")
        return

    # Create visualization figure
    fig, axs = plt.subplots(6, 1, figsize=(6, 18))
    
    try:
        print(f"\nProcessing {os.path.basename(csv_file)}...")
        phase_info, df = analyze_fermentation_phases(csv_file)

        # Specify the single batch base name
        batch_base_name = os.path.basename(csv_file).replace(' MF.csv', '')

        # Load sensor data from the pickle file
        pkl_file = f'{batch_base_name} wiffs.pkl'
        if os.path.exists(pkl_file):
            with open(pkl_file, 'rb') as f:
                wiffs = pickle.load(f)
            
            # Extract timestamps and sensor data
            timestamps = [entry['timestamp'] for entry in wiffs]
            timestamps = np.array(timestamps)
            sorted_indices = np.argsort(timestamps)
            timestamps = timestamps[sorted_indices]
            timestamps = [datetime.fromtimestamp(ts) for ts in timestamps]
            
            data = np.array([entry['data'] for entry in wiffs])
            data = data[sorted_indices]
            
            # Process each sensor (assuming 4 sensors, each with 128 channels)
            sensors = []
            for i in range(4):
                sensor_data = data[:, i, :]
                sensor_mean = np.mean(sensor_data, axis=1)
                sensors.append(sensor_mean)
            
            # Convert timestamps to hours from start
            start_time = timestamps[0]
            hours = [(ts - start_time).total_seconds() / 3600 for ts in timestamps]

            # Normalize hours to start at 0.0
            hours = [h - hours[0] for h in hours]

            # Plot sensor data for each of the four sensors with smoothed line
            for i in range(4):
                axs[i+2].plot(hours, sensors[i], label=f'Sensor {i+1} Data', alpha=0.5)
                # Apply Savitzky-Golay filter for smoothing
                smoothed = savgol_filter(sensors[i], window_length=11, polyorder=2)
                axs[i+2].plot(hours, smoothed, label=f'Sensor {i+1} Smoothed', color='blue',alpha=1)
                axs[i+2].set_title(f'Sensor {i+1} Data')
                axs[i+2].set_xlabel('Time (hours)')
                axs[i+2].set_ylabel('Sensor Value')
                axs[i+2].set_xlim(left=0)
                axs[i+2].set_ylim(bottom=0)
                axs[i+2].set_ylim(top=max(sensors[i])*1.1)
                axs[i+2].legend(loc='upper right')
                axs[i+2].grid(True)
        else:
            print(f"Sensor data file {pkl_file} not found.")

        # Calculate smoothed data using Savitzky-Golay filter
        df['smoothed_data'] = savgol_filter(df['mass flow (CO2)'], window_length=11, polyorder=2)

        # Insert the smoothed data column after the 'mass flow (CO2)' column
        columns = list(df.columns)
        co2_index = columns.index('mass flow (CO2)')
        columns.insert(co2_index + 1, 'smoothed_data')
        df = df[columns]

        # Plot mass flow data and phases
        axs[0].plot(df.index, df['mass flow (CO2)'], label='Raw Data', alpha=0.5)
        axs[0].plot(df.index, df['smoothed'], label='Smoothed', color='blue', alpha=1.00)
        axs[0].set_xlim(0, df.index.max())  # Set x-axis limit to max time value
        axs[0].set_xticks(np.arange(0, df.index.max() + 10, 10))  # Set x-ticks every 10 hours for mass flow data

        colors = {
            'Lag Phase': 'blue',
            'Exponential Phase': 'orange',
            'Production Phase': 'green',
            'Death Phase': 'gray',
            'Post Death Phase': 'black'
        }
        for phase in phase_info:
            axs[0].axvspan(phase['Start Hour'], phase['End Hour'], 
                          alpha=0.2, label=phase['Phase'], 
                          color=colors[phase['Phase']])

        axs[0].set_title(batch_base_name)

        axs[0].set_xlabel('Time (hours)')
        axs[0].set_ylabel('Mass Flow (CO2)')
        axs[0].legend(loc='upper right', fontsize='small')
        axs[0].grid(True)

        # Plot derivatives
        axs[1].plot(df.index, df['first_derivative'], label='First Derivative', alpha=0.5)
        axs[1].plot(df.index, df['second_derivative'], label='Second Derivative',color='red', alpha=0.5)
        axs[1].plot(df.index, df['rolling_derivative'], label='Smoothed Derivative', color='blue', alpha=1)
        axs[1].set_xlabel('Time (hours)')
        axs[1].set_ylabel('Rate of Change')
        axs[1].set_xlim(0, df.index.max())  # Set x-axis limit to max time value for derivatives
        axs[1].set_xticks(np.arange(0, df.index.max() + 10, 10))  # Set x-ticks every 10 hours for derivatives
        axs[1].set_ylim(df['first_derivative'].min(), df['first_derivative'].max())  # Set y-axis limit to min and max derivative values
        axs[1].legend(loc='upper right', fontsize='small')
        axs[1].grid(True)

        # Prepare data for Excel
        excel_data = []
        for idx, row in df.iterrows():
            # Find the phase label for the current index
            phase_label = next((phase['Phase'] for phase in phase_info if phase['Start Hour'] <= idx <= phase['End Hour']), 'Unknown')
            phase_label = phase_label.replace(' Phase', '')
            idx_minutes = idx * 60  # Convert hours to minutes
            # Ensure smoothed_data is a single numeric value
            smoothed_value = row['smoothed_data'] if isinstance(row['smoothed_data'], (int, float)) else row['smoothed_data'].iloc[0]
            excel_data.append([
                idx,  # Timestamp in hours
                idx_minutes,  # Timestamp in minutes
                row['mass flow (CO2)'],  # CO2 mass flow
                smoothed_value,  # Smoothed data
                row['first_derivative'],  # 1st derivative
                row['second_derivative'],  # 2nd derivative
                row['rolling_derivative'],  # Smoothed derivative
                phase_label  # Phase label
            ])
            # Add sensor data
            for sensor in sensors:
                excel_data[-1].append(sensor[int(idx * len(sensor) / len(df))])

        # Write data to Excel
        wb = Workbook()
        ws = wb.active
        ws.title = 'Fermentation Data'
        headers = ['Timestamp (hours)', 'Timestamp (minutes)', 'CO2 Mass Flow (sccm)', 'Smoothed Data', '1st Derivative', '2nd Derivative', 'Smoothed Derivative', 'Phase']
        headers.extend([f'Sensor {i+1}' for i in range(4)])
        ws.append(headers)
        for row in excel_data:
            ws.append(row)

        # Apply comma style with three decimal places to numeric columns
        comma_style = NamedStyle(name='comma_style', number_format='#,##0.000')
        for col in range(1, len(headers)):
            for cell in ws.iter_cols(min_col=col+1, max_col=col+1, min_row=2):
                for c in cell:
                    c.style = comma_style

        # Save the Excel file
        script_directory = os.path.dirname(os.path.abspath(__file__))
        excel_path = os.path.join(script_directory, f'{batch_base_name}_fermentation_data.xlsx')
        wb.save(excel_path)

        # Load the Excel file
        data = pd.read_excel(excel_path)

        # Display the first few rows to understand the structure
        # print(data.head())

    except Exception as e:
        print(f"Error processing {csv_file}: {str(e)}")

    # Adjust figure layout
    plt.tight_layout()

    # Change the figure path to be where the script is located
    script_directory = os.path.dirname(os.path.abspath(__file__))
    figure_path = os.path.join(script_directory, f'{batch_base_name}_MF&Wiffs.png')
    plt.savefig(figure_path)
    plt.show()

    return fig

if __name__ == "__main__":
    # Set the working directory
    os.chdir(r"C:\Users\GeorgiosBalamotis\sqale.ai\3-PRODUCT DEVELOPMENT - PROJECT-101-eNose - PROJECT-101-eNose\2024 12 16-Experimental data generated\Fermentations_DL\Datasets")
    
    # Specify the single batch base name
    batch_base_name = 'Batch7'  # Replace with the desired base name
    
    # Construct file names
    single_batch_file = f'{batch_base_name} MF.csv'
    pkl_file = f'{batch_base_name} wiffs.pkl'
    analyze_single_batch(single_batch_file)
