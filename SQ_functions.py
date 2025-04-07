import os
import glob
import webbrowser
import re
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import plotly.graph_objs as go
import plotly.express as px
import plotly.io as pio
from tqdm import tqdm
from sklearn.linear_model import LinearRegression
from sklearn.preprocessing import PolynomialFeatures
from sklearn.metrics import mean_squared_error, mean_absolute_error, r2_score
from scipy.optimize import minimize
from scipy.optimize import least_squares
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.formatting.rule import CellIsRule
from datetime import datetime

# --- GENERATED FUNCTIONS ---
# --- 1. DATA PREPERATION ---
# get_datasets - Collects and categorizes CSV files
# merge_sensor_pairs - Merges sensor data into one dataset along with extra columns
# assign_real_values - Assigns real values to a DataFrame based on time ranges

def get_datasets(csv_directory):
    sfmmf_co2_files = glob.glob(os.path.join(csv_directory, 'SFMMF_eNc_CO2_SLPM_*.csv'))
    alicat_air_files = glob.glob(os.path.join(csv_directory, 'Alicat_Air_*.csv'))
    sfmmf_air_files = glob.glob(os.path.join(csv_directory, 'SFMMF_eNc_Air_SLPM_*.csv'))
    alicat_co2_files = glob.glob(os.path.join(csv_directory, 'Alicat_CO2_*.csv'))
    sfmmf_both_varying_files = glob.glob(os.path.join(csv_directory, 'SFMMF_eNc_BV_*.csv'))
    alicat_bv_air_files = glob.glob(os.path.join(csv_directory, 'Alicat_BV_Air_*.csv'))
    alicat_bv_co2_files = glob.glob(os.path.join(csv_directory, 'Alicat_BV_CO2_*.csv'))

    # Count the number of datasets read
    sfmmf_co2_count = len(sfmmf_co2_files)
    alicat_air_count = len(alicat_air_files)
    sfmmf_air_count = len(sfmmf_air_files)
    alicat_co2_count = len(alicat_co2_files)
    sfmmf_bv_count = len(sfmmf_both_varying_files)
    alicat_bv_air_count = len(alicat_bv_air_files)
    alicat_bv_co2_count = len(alicat_bv_co2_files)

    return (sfmmf_co2_files, alicat_air_files, sfmmf_air_files, alicat_co2_files, 
            sfmmf_both_varying_files, alicat_bv_air_files, alicat_bv_co2_files,
            sfmmf_co2_count, alicat_air_count, sfmmf_air_count, alicat_co2_count, 
            sfmmf_bv_count, alicat_bv_air_count, alicat_bv_co2_count)

def merge_sensor_pairs(sfmmf_co2_files, alicat_air_files,
                       sfmmf_air_files, alicat_co2_files,
                       sfmmf_bv_files, alicat_bv_air_files, alicat_bv_co2_files):
    
    all_merged_data = pd.DataFrame()

    # --- 1. CO2 fixed ---
    for sfmmf_file, alicat_file in zip(sfmmf_co2_files, alicat_air_files):
        co2_value = float(re.search(r'\d+\.\d+', sfmmf_file).group())

        sfmmf_df = pd.read_csv(sfmmf_file, delimiter=';')
        sfmmf_df.columns = sfmmf_df.columns.str.strip()
        alicat_df = pd.read_csv(alicat_file, delimiter=';')
        alicat_df.columns = alicat_df.columns.str.strip()

        sfmmf_df.rename(columns={'CO2 conc.': 'CO2_Concentration', 'Mass flow': 'Mass_Flow', 'Sensirion_Mass_Flow': 'Mass_Flow'}, inplace=True)
        alicat_df.rename(columns={'alicat_flowrate_t_avg': 'Alicat_Air_PMF', 'Alicat_Air_PMF_Average': 'Alicat_Air_PMF'}, inplace=True)

        sfmmf_df['CO2_Concentration'] = sfmmf_df['CO2_Concentration'] / 100
        sfmmf_df['Timestamp'] = pd.to_datetime(sfmmf_df['Timestamp'])
        alicat_df['Air PMF [SLPM]'] = alicat_df['Alicat_Air_PMF']
        alicat_df['Timestamp'] = pd.to_datetime(alicat_df['Timestamp'])

        merged = pd.merge(sfmmf_df, alicat_df[['Timestamp', 'Air PMF [SLPM]']], on='Timestamp', how='inner')
        merged['CO2 PMF [SLPM]'] = co2_value
        all_merged_data = pd.concat([all_merged_data, merged], ignore_index=True)

    # --- 2. Air fixed ---
    for sfmmf_file, alicat_file in zip(sfmmf_air_files, alicat_co2_files):
        air_value = float(re.search(r'\d+\.\d+', sfmmf_file).group())

        sfmmf_df = pd.read_csv(sfmmf_file, delimiter=';')
        sfmmf_df.columns = sfmmf_df.columns.str.strip()
        alicat_df = pd.read_csv(alicat_file, delimiter=';')
        alicat_df.columns = alicat_df.columns.str.strip()

        sfmmf_df.rename(columns={'CO2 conc.': 'CO2_Concentration', 'Mass flow': 'Mass_Flow', 'Sensirion_Mass_Flow': 'Mass_Flow'}, inplace=True)
        alicat_df.rename(columns={'alicat_flowrate_t_avg 2': 'Alicat_CO2_PMF', 'Alicat_CO2_PMF_Average': 'Alicat_CO2_PMF'}, inplace=True)

        sfmmf_df['CO2_Concentration'] = sfmmf_df['CO2_Concentration'] / 100
        sfmmf_df['Timestamp'] = pd.to_datetime(sfmmf_df['Timestamp'])
        alicat_df['CO2 PMF [SLPM]'] = alicat_df['Alicat_CO2_PMF']
        alicat_df['Timestamp'] = pd.to_datetime(alicat_df['Timestamp'])

        merged = pd.merge(sfmmf_df, alicat_df[['Timestamp', 'CO2 PMF [SLPM]']], on='Timestamp', how='inner')
        merged['Air PMF [SLPM]'] = air_value
        all_merged_data = pd.concat([all_merged_data, merged], ignore_index=True)

    # --- 3. Both Varying ---
    for sfmmf_file, alicat_air_file, alicat_co2_file in zip(sfmmf_bv_files, alicat_bv_air_files, alicat_bv_co2_files):
        sfmmf_df = pd.read_csv(sfmmf_file, delimiter=';')
        sfmmf_df.columns = sfmmf_df.columns.str.strip()
        alicat_air_df = pd.read_csv(alicat_air_file, delimiter=';')
        alicat_air_df.columns = alicat_air_df.columns.str.strip()
        alicat_co2_df = pd.read_csv(alicat_co2_file, delimiter=';')
        alicat_co2_df.columns = alicat_co2_df.columns.str.strip()

        sfmmf_df.rename(columns={'CO2 conc.': 'CO2_Concentration', 'Mass flow': 'Mass_Flow', 'Sensirion_Mass_Flow': 'Mass_Flow'}, inplace=True)
        alicat_air_df.rename(columns={'alicat_flowrate_t_avg': 'Alicat_Air_PMF', 'Alicat_Air_PMF_Average': 'Alicat_Air_PMF'}, inplace=True)
        alicat_co2_df.rename(columns={'alicat_flowrate_t_avg 2': 'Alicat_CO2_PMF', 'Alicat_CO2_PMF_Average': 'Alicat_CO2_PMF'}, inplace=True)

        sfmmf_df['CO2_Concentration'] = sfmmf_df['CO2_Concentration'] / 100
        sfmmf_df['Timestamp'] = pd.to_datetime(sfmmf_df['Timestamp'])
        alicat_air_df['Air PMF [SLPM]'] = alicat_air_df['Alicat_Air_PMF']
        alicat_air_df['Timestamp'] = pd.to_datetime(alicat_air_df['Timestamp'])
        alicat_co2_df['CO2 PMF [SLPM]'] = alicat_co2_df['Alicat_CO2_PMF']
        alicat_co2_df['Timestamp'] = pd.to_datetime(alicat_co2_df['Timestamp'])

        merged = pd.merge(sfmmf_df, alicat_air_df[['Timestamp', 'Air PMF [SLPM]']], on='Timestamp', how='inner')
        merged = pd.merge(merged, alicat_co2_df[['Timestamp', 'CO2 PMF [SLPM]']], on='Timestamp', how='inner')
        all_merged_data = pd.concat([all_merged_data, merged], ignore_index=True)

    return all_merged_data

def assign_real_values(df, real_values, column_name):
    df[column_name] = None
    for i in range(len(real_values) - 1):
        start_time, value = real_values[i]
        end_time = real_values[i + 1][0]

        df.loc[
            (df["Timestamp"] >= start_time) & (df["Timestamp"] < end_time),
            column_name
        ] = value

    # Assign last range value
    df.loc[df["Timestamp"] >= real_values[-1][0], column_name] = real_values[-1][1]

# --- 2. CORE FUNCTIONS --- 
# adjusted_r2 - Calculates the adjusted R² value
# constrained_polynomial_fit_massbalance - Fits a polynomial model with MF constraint CO₂ + Air ≈ total mass flow
# constrained_polynomial_fit_massbalance_nonneg - Fits a polynomial model with MF constraint CO₂ + Air ≈ total mass flow and non-negativity constraints
# predict_constrained - Uses fitted constrained weights and polynomial basis to predict CO₂ and Air flows
# evaluate_constrained_fit - Evaluates the constrained polynomial fit
# train_unconstrained_models - Trains unconstrained polynomial models for CO₂ and Air

def adjusted_r2(y_true, y_pred, n_features):
    r2 = r2_score(y_true, y_pred)
    n = len(y_true)
    return 1 - (1 - r2) * (n - 1) / (n - n_features - 1)

def constrained_polynomial_fit_massbalance(X, Y_co2, Y_air, degree, constraint_weight=1.0):
    if X.ndim != 2 or X.shape[1] != 2:
        raise ValueError(f"Expected X shape (n_samples, 2), got {X.shape}")

    mask = (~np.isnan(X).any(axis=1)) & (~np.isinf(X).any(axis=1)) \
         & (~np.isnan(Y_co2)) & (~np.isinf(Y_co2)) \
         & (~np.isnan(Y_air)) & (~np.isinf(Y_air))

    X_clean = X[mask]
    Y_co2_clean = Y_co2[mask]
    Y_air_clean = Y_air[mask]

    if len(X_clean) == 0:
        raise ValueError("No valid data points after cleaning.")

    poly = PolynomialFeatures(degree=degree, include_bias=True)
    Phi = poly.fit_transform(X_clean)
    n_features = Phi.shape[1]

    def unpack_weights(w):
        return w[:n_features], w[n_features:]

    def residuals(w):
        w_z, w_k = unpack_weights(w)
        z_pred = Phi @ w_z
        k_pred = Phi @ w_k
        r1 = Y_co2_clean - z_pred
        r2 = Y_air_clean - k_pred
        r3 = constraint_weight * ((z_pred + k_pred) - X_clean[:, 0])
        return np.concatenate([r1, r2, r3])

    w0 = np.zeros(2 * n_features)
    result = least_squares(residuals, w0, method='trf')

    if not result.success:
        raise RuntimeError("Least squares optimization failed: " + result.message)

    w_z, w_k = unpack_weights(result.x)
    return w_z, w_k, poly

def constrained_polynomial_fit_massbalance_nonneg(X, Y_co2, Y_air, degree, constraint_weight=1.0):
    if X.ndim != 2 or X.shape[1] != 2:
        raise ValueError(f"Expected X shape (n_samples, 2), got {X.shape}")

    mask = (~np.isnan(X).any(axis=1)) & (~np.isinf(X).any(axis=1)) \
         & (~np.isnan(Y_co2)) & (~np.isinf(Y_co2)) \
         & (~np.isnan(Y_air)) & (~np.isinf(Y_air))

    X_clean = X[mask]
    Y_co2_clean = Y_co2[mask]
    Y_air_clean = Y_air[mask]

    if len(X_clean) == 0:
        raise ValueError("No valid data points after cleaning.")

    poly = PolynomialFeatures(degree=degree, include_bias=True)
    Phi = poly.fit_transform(X_clean)
    n_features = Phi.shape[1]

    def unpack(w):
        return w[:n_features], w[n_features:]

    def loss(w):
        w_z, w_k = unpack(w)
        Z_pred = Phi @ w_z
        K_pred = Phi @ w_k
        mse_z = np.mean((Y_co2_clean - Z_pred)**2)
        mse_k = np.mean((Y_air_clean - K_pred)**2)
        mse_bal = np.mean(((Z_pred + K_pred) - X_clean[:, 0])**2)
        return mse_z + mse_k + constraint_weight * mse_bal

    def nonneg_constraint(w):
        w_z, w_k = unpack(w)
        return np.concatenate([Phi @ w_z, Phi @ w_k])

    constraints = {'type': 'ineq', 'fun': nonneg_constraint}
    w0 = np.zeros(2 * n_features)

    # Progress bar setup
    pbar = tqdm(total=5000, desc="Optimizing (nonneg)", position=0)
    iteration = [0]  # mutable object to track step

    def callback(w, state):
        iteration[0] += 1
        pbar.update(1)
        if iteration[0] >= 5000:
            pbar.close()

    result = minimize(
        loss,
        w0,
        method='trust-constr',
        constraints=[constraints],
        callback=callback,
        options={'maxiter': 5000, 'gtol': 1e-6, 'verbose': 0}
    )

    pbar.close()

    if not result.success:
        raise RuntimeError("Optimization failed: " + result.message)

    w_z, w_k = unpack(result.x)
    return w_z, w_k, poly

def predict_constrained(X, w_co2, w_air, poly):
    Phi = poly.transform(X)
    co2_pred = Phi @ w_co2
    air_pred = Phi @ w_air
    return co2_pred, air_pred

def evaluate_constrained_fit(X, Y_co2, Y_air, w_co2, w_air, poly, degree=None):
    Phi = poly.transform(X)
    n_features = Phi.shape[1]

    z_pred = Phi @ w_co2
    k_pred = Phi @ w_air

    z_rmse = np.sqrt(mean_squared_error(Y_co2, z_pred))
    z_r2 = r2_score(Y_co2, z_pred)
    z_adj_r2 = adjusted_r2(Y_co2, z_pred, n_features)
    z_mse = mean_squared_error(Y_co2, z_pred)
    z_mae = mean_absolute_error(Y_co2, z_pred)

    k_rmse = np.sqrt(mean_squared_error(Y_air, k_pred))
    k_r2 = r2_score(Y_air, k_pred)
    k_adj_r2 = adjusted_r2(Y_air, k_pred, n_features)
    k_mse = mean_squared_error(Y_air, k_pred)
    k_mae = mean_absolute_error(Y_air, k_pred)

    constraint_viol = (z_pred + k_pred) - X[:, 0]
    max_violation = np.max(np.abs(constraint_viol))

    return {
        "Degree": degree,
        "CO2_RMSE": z_rmse, "CO2_MSE": z_mse, "CO2_MAE": z_mae, "CO2_R2": z_r2, "CO2_Adj_R2": z_adj_r2,
        "Air_RMSE": k_rmse, "Air_MSE": k_mse, "Air_MAE": k_mae, "Air_R2": k_r2, "Air_Adj_R2": k_adj_r2,
        "Max_Constraint_Error": max_violation,
        "CO2_Pred": z_pred,
        "Air_Pred": k_pred,
        "Constraint_Violation": constraint_viol  # Return the actual array
    }

def train_unconstrained_models(df, feature_names, min_order=1, max_order=4):
    from sklearn.linear_model import LinearRegression
    from sklearn.preprocessing import PolynomialFeatures
    from sklearn.metrics import mean_squared_error, r2_score, mean_absolute_error

    models = {"CO2": {}, "Air": {}}
    polynomials = {"CO2": {}, "Air": {}}
    metrics = {"CO2": {}, "Air": {}}
    columns_order = list(df.columns)

    for degree in range(min_order, max_order + 1):
        for gas in ["CO2", "Air"]:
            target_col = f"{gas} PMF [SLPM]"
            poly = PolynomialFeatures(degree=degree, include_bias=True)
            X_poly = poly.fit_transform(df[feature_names].values)
            model = LinearRegression().fit(X_poly, df[target_col].values)
            y_pred = model.predict(X_poly)

            models[gas][degree] = model
            polynomials[gas][degree] = poly

            mse = mean_squared_error(df[target_col].values, y_pred)
            rmse = np.sqrt(mse)
            mae = mean_absolute_error(df[target_col].values, y_pred)
            r2 = r2_score(df[target_col].values, y_pred)
            adj_r2 = 1 - ((1 - r2) * (len(y_pred) - 1) / (len(y_pred) - X_poly.shape[1] - 1))

            metrics[gas][degree] = {
                "MSE": mse, "RMSE": rmse, "MAE": mae,
                "R2": r2, "Adj_R2": adj_r2
            }

            df[f"{gas}_PMF_Pred_Deg{degree}"] = y_pred
            df[f"{gas}_Pred_Diff_Deg{degree}"] = (y_pred - df[target_col]) / df[target_col]
            columns_order.extend([
                f"{gas}_PMF_Pred_Deg{degree}",
                f"{gas}_Pred_Diff_Deg{degree}"
            ])

    df = df[columns_order]
    return df, models, polynomials, metrics

# --- 3. VISUALIZATION FUNCTIONS ---
# plot_constrained_surface - Plots the constrained polynomial surface
# plot_unconstrained_surfaces - Plots the unconstrained polynomial surfaces
# plot_deviation_heatmap - Plots the deviation heatmap
# plot_mass_balance_violation - Plots histogram of mass balance violation

def plot_constrained_surface(w, poly, label, degree, outdir, X_real, Y_real):  
    # Grid range
    mf_min, mf_max = np.min(X_real[:, 0]), np.max(X_real[:, 0])
    co2_min, co2_max = np.min(X_real[:, 1]), np.max(X_real[:, 1])

    grid_x, grid_y = np.meshgrid(
        np.linspace(mf_min, mf_max, 50),
        np.linspace(co2_min, co2_max, 50)
    )

    grid_input = np.column_stack([grid_x.ravel(), grid_y.ravel()])
    Phi = poly.transform(grid_input)
    z_grid = Phi @ w
    z_grid = z_grid.reshape(grid_x.shape)

    # Surface
    surface = go.Surface(
        x=grid_x, y=grid_y, z=z_grid,
        colorscale='Viridis',
        opacity=0.7,
        name='Predicted Surface'
    )

    # Data points
    scatter = go.Scatter3d(
        x=X_real[:, 0], y=X_real[:, 1], z=Y_real,
        mode='markers',
        marker=dict(size=3, color='blue'),
        name='Actual Data'
    )

    layout = go.Layout(
        title=f"{label} PMF - Degree {degree}",
        scene=dict(
            xaxis_title="Total Mass Flow [SLPM]",
            yaxis_title="CO2 Concentration",
            zaxis_title=f"{label} PMF [SLPM]"
        ),
        margin=dict(l=0, r=0, b=0, t=40)
    )

    fig = go.Figure(data=[surface, scatter], layout=layout)
    filename = os.path.join(outdir, f"surface_{label}_deg{degree}.html")
    fig.show()

def plot_unconstrained_surfaces(df, models, polynomials, min_order, max_order):
    mass_flow_range = np.linspace(0, 25, 100)
    co2_conc_range = np.linspace(0, 0.5, 100)
    mass_flow_grid, co2_conc_grid = np.meshgrid(mass_flow_range, co2_conc_range)
    grid_points = np.c_[mass_flow_grid.ravel(), co2_conc_grid.ravel()]

    for degree in range(min_order, max_order + 1):
        for gas in ["CO2", "Air"]:
            poly = polynomials[gas][degree]
            model = models[gas][degree]
            Z_actual = df[f"{gas} PMF [SLPM]"].values
            Z_surface = model.predict(poly.transform(grid_points)).reshape(mass_flow_grid.shape)
            Z_surface = np.clip(Z_surface, a_min=0, a_max=None)

            fig = go.Figure()
            fig.add_trace(go.Scatter3d(
                x=df["Mass_Flow"], y=df["CO2_Concentration"], z=Z_actual,
                mode='markers', marker=dict(size=5, color='blue'), name='Data'
            ))

            fig.add_trace(go.Surface(
                x=mass_flow_range, y=co2_conc_range, z=Z_surface,
                colorscale='Viridis', opacity=0.7, name=f'Degree {degree}'
            ))

            fig.update_layout(
                title=f"{gas} Polynomial Regression Degree {degree}",
                scene=dict(
                    xaxis_title='Mass flow',
                    yaxis_title='CO2 conc.',
                    zaxis_title=f'{gas} PMF [SLPM]'
                )
            )
            fig.show()

def plot_deviation_heatmap(df, diff_column, title):

    if diff_column not in df.columns or df[diff_column].dropna().empty:
        print(f"Skipping {diff_column}: column not found or all values are NaN.")
        return
    fig = px.density_heatmap(
        df, x="Mass_Flow", y="CO2_Concentration", z=diff_column,
        nbinsx=40, nbinsy=40, color_continuous_scale="RdBu",
        title=title, histfunc="avg"
    )
    fig.update_layout(
        xaxis_title="Mass Flow [SLPM]",
        yaxis_title="CO2 Concentration",
        coloraxis_colorbar=dict(title="% Deviation")
    )
    fig.show()

def plot_mass_balance_violation(df, min_order, max_order):
    """
    Displays interactive Plotly histograms of mass balance violation 
    (CO2_Pred + Air_Pred - Mass_Flow) for each polynomial degree.
    """
    for degree in range(min_order, max_order + 1):
        co2_col = f"CO2_PMF_Pred_Deg{degree}"
        air_col = f"Air_PMF_Pred_Deg{degree}"

        if co2_col not in df.columns or air_col not in df.columns:
            print(f"Skipping degree {degree}: predictions not found.")
            continue

        df[f"Mass_Balance_Violation_Deg{degree}"] = (
            df[co2_col] + df[air_col] - df["Mass_Flow"]
        )

        violations = df[f"Mass_Balance_Violation_Deg{degree}"].dropna()
        if violations.empty:
            print(f"No valid data for degree {degree}.")
            continue

        fig = px.histogram(
            violations,
            nbins=50,
            title=f"Mass Balance Violation - Degree {degree}",
            labels={'value': 'Violation [SLPM]'},
            opacity=0.75
        )
        fig.update_layout(
            xaxis_title="Violation [SLPM]",
            yaxis_title="Count",
            bargap=0.1
        )
        fig.show()

# --- 4. EXPORTS ---
# export_constrained_equations - Exports the polynomial equations and metrics to a text file
# export_to_excel - Exports dataframes to an Excel file

def export_constrained_equations(w_co2, w_air, poly, output_path, degree=None, metrics=None, append=False):
    powers = poly.powers_
    mode = 'a' if append else 'w'

    with open(output_path, mode) as f:
        # Write intro only on first write
        if not append:
            f.write("For Z: CO2 Partial Mass Flow [SLPM] and K: Air Partial Mass Flow [SLPM], "
                    "X: Mass flow from SFM [SLPM], and Y: CO2 conc. from eNose [%], the polynomial equations are:\n")

        f.write(f"\n--- Degree {degree} ---\n")

        # CO2
        f.write("CO2 Partial Mass Flow Equation:\n")
        f.write(format_polynomial_equation(w_co2, powers, 'Z') + "\n")
        if metrics:
            m = metrics
            f.write(f"R2: {m['CO2_R2']:.4f}, Adjusted R2: {m['CO2_Adj_R2']:.4f}, "
                    f"MSE: {m['CO2_MSE']:.4f}, RMSE: {m['CO2_RMSE']:.4f}, MAE: {m['CO2_MAE']:.4f}\n\n")

        # Air
        f.write("Air Partial Mass Flow Equation:\n")
        f.write(format_polynomial_equation(w_air, powers, 'K') + "\n")
        if metrics:
            f.write(f"R2: {m['Air_R2']:.4f}, Adjusted R2: {m['Air_Adj_R2']:.4f}, "
                    f"MSE: {m['Air_MSE']:.4f}, RMSE: {m['Air_RMSE']:.4f}, MAE: {m['Air_MAE']:.4f}\n\n")

        # Constraint Violation
        if metrics:
            f.write(f"Max Constraint Violation: {m['Max_Constraint_Error']:.6f} SLPM\n")

def export_unconstrained_equations(models, polynomials, metrics, feature_names, output_path):
    mode = 'w'
    with open(output_path, mode) as f:
        f.write("For Z: CO2 PMF [SLPM] and K: Air PMF [SLPM], X: Gas MF from SFM [SLPM], and Y: CO2 conc. from eNose [%] we have:\n")

        degrees = sorted(models["CO2"].keys())

        for degree in degrees:
            f.write(f"\n--- Degree {degree} ---\n")

            # --- CO2 ---
            poly_co2 = polynomials['CO2'][degree]
            coefs_co2 = models['CO2'][degree].coef_
            intercept_co2 = models['CO2'][degree].intercept_
            powers_co2 = poly_co2.powers_
            eq_co2 = format_polynomial_equation(coefs_co2, powers_co2, 'Z') + f" + ({intercept_co2:.4f})"

            m = metrics['CO2'][degree]
            f.write("CO2 Partial Mass Flow Equation:\n")
            f.write(eq_co2 + "\n")
            f.write(f"R2: {m['R2']:.4f}, Adjusted R2: {m['Adj_R2']:.4f}, "
                    f"MSE: {m['MSE']:.4f}, RMSE: {m['RMSE']:.4f}, MAE: {m['MAE']:.4f}\n\n")

            # --- Air ---
            poly_air = polynomials['Air'][degree]
            coefs_air = models['Air'][degree].coef_
            intercept_air = models['Air'][degree].intercept_
            powers_air = poly_air.powers_
            eq_air = format_polynomial_equation(coefs_air, powers_air, 'K') + f" + ({intercept_air:.4f})"

            m = metrics['Air'][degree]
            f.write("Air Partial Mass Flow Equation:\n")
            f.write(eq_air + "\n")
            f.write(f"R2: {m['R2']:.4f}, Adjusted R2: {m['Adj_R2']:.4f}, "
                    f"MSE: {m['MSE']:.4f}, RMSE: {m['RMSE']:.4f}, MAE: {m['MAE']:.4f}\n")

def export_to_excel(path, sheets_dict):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    mode = 'a' if os.path.exists(path) else 'w'
    
    if mode == 'a':
        with pd.ExcelWriter(path, engine='openpyxl', mode=mode, if_sheet_exists='replace') as writer:
            for sheet_name, df in sheets_dict.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)
    else:
        with pd.ExcelWriter(path, engine='openpyxl', mode=mode) as writer:
            for sheet_name, df in sheets_dict.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)

# --- 5. MISC FUNCTIONS ---
# score - Calculates a score based on adjusted R² and polynomial degree
# stress_test - Runs a stress test for Unconstraint case on the models and saves the results
# stress_test_constrained - Runs a stress test for Constrained case on the models and saves the results

def score(degree, adj_r2, weight_adj_r2=0.95, weight_degree=0.05):
    max_degree = max(models["CO2"].keys())
    normalized_degree = degree / max_degree
    return (weight_adj_r2 * adj_r2) - (weight_degree * normalized_degree)

def stress_test(models, polynomials, output_dir, min_order=1, max_order=4, current_date=None):
    if current_date is None:
        current_date = datetime.now().strftime('%Y-%m-%d')
    
    print("Running Stress Test...")

    X_values = np.linspace(0.50, 10.00, num=int((10.00 - 0.50) / 0.5) + 1)
    Y_values = np.linspace(0.05, 0.35, num=int((0.35 - 0.05) / 0.025) + 1)

    results_dict = {
        "Mass_Flow": [],
        "CO2_Concentration": [],
        "CO2_Th_PMF": [],
        "Air_Th_PMF": []
    }

    for degree in range(min_order, max_order+1):
        results_dict[f"CO2_Pred_Deg{degree}"] = []
        results_dict[f"Air_Pred_Deg{degree}"] = []

    for X in X_values:
        for Y in Y_values:
            results_dict["Mass_Flow"].append(X)
            results_dict["CO2_Concentration"].append(Y)
            results_dict["CO2_Th_PMF"].append(X * Y)
            results_dict["Air_Th_PMF"].append(X * (1 - Y))

            for degree in range(min_order, max_order+1):
                poly = polynomials["CO2"][degree]
                X_transformed = poly.transform([[X, Y]])
                co2_pred = models["CO2"][degree].predict(X_transformed)[0]
                air_pred = models["Air"][degree].predict(X_transformed)[0]
                results_dict[f"CO2_Pred_Deg{degree}"].append(co2_pred)
                results_dict[f"Air_Pred_Deg{degree}"].append(air_pred)

    df_st = pd.DataFrame(results_dict)

    co2_pred_columns = [col for col in df_st.columns if "CO2_Pred_Deg" in col]
    air_pred_columns = [col for col in df_st.columns if "Air_Pred_Deg" in col]

    df_st["CO2_Best_guess"] = df_st[co2_pred_columns].apply(
        lambda row: row.dropna().loc[(row.dropna() - df_st.loc[row.name, "CO2_Th_PMF"]).abs().idxmin()]
        if not row.dropna().empty else np.nan,
        axis=1
    )
    df_st["Air_Best_guess"] = df_st[air_pred_columns].apply(
        lambda row: row.dropna().loc[(row.dropna() - df_st.loc[row.name, "Air_Th_PMF"]).abs().idxmin()]
        if not row.dropna().empty else np.nan,
        axis=1
    )

    ordered_columns = ["Mass_Flow", "CO2_Concentration", "CO2_Th_PMF", "CO2_Best_guess"] +                       co2_pred_columns + ["Air_Th_PMF", "Air_Best_guess"] + air_pred_columns
    df_st = df_st[ordered_columns]

    excel_path = os.path.join(output_dir, f'Data_Analysis_{current_date}.xlsx')
    mode = 'a' if os.path.exists(excel_path) else 'w'
    if mode == 'a':
        with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df_st.to_excel(writer, sheet_name='Stress_Test', index=False)
    else:
        with pd.ExcelWriter(excel_path, engine='openpyxl', mode='w') as writer:
            df_st.to_excel(writer, sheet_name='Stress_Test', index=False)


    wb = load_workbook(excel_path)
    ws = wb["Stress_Test"]
    red_font = Font(color="FF0000")
    green_font = Font(color="008000")

    for col in range(3, len(df_st.columns) + 2):
        col_letter = ws.cell(row=1, column=col).column_letter
        ws.conditional_formatting.add(
            f"{col_letter}2:{col_letter}{len(df_st) + 1}",
            CellIsRule(operator="lessThan", formula=["0"], stopIfTrue=True, font=red_font),
        )

    for row_idx in range(2, len(df_st) + 2):
        co2_best_value = df_st.at[row_idx - 2, "CO2_Best_guess"]
        air_best_value = df_st.at[row_idx - 2, "Air_Best_guess"]

        for col_name in co2_pred_columns:
            if df_st.at[row_idx - 2, col_name] == co2_best_value:
                col_letter = ws.cell(row=1, column=df_st.columns.get_loc(col_name) + 1).column_letter
                ws.conditional_formatting.add(
                    f"{col_letter}{row_idx}:{col_letter}{row_idx}",
                    CellIsRule(operator="greaterThanOrEqual", formula=["0"], stopIfTrue=True, font=green_font),
                )
                break

        for col_name in air_pred_columns:
            if df_st.at[row_idx - 2, col_name] == air_best_value:
                col_letter = ws.cell(row=1, column=df_st.columns.get_loc(col_name) + 1).column_letter
                ws.conditional_formatting.add(
                    f"{col_letter}{row_idx}:{col_letter}{row_idx}",
                    CellIsRule(operator="greaterThanOrEqual", formula=["0"], stopIfTrue=True, font=green_font),
                )
                break

    wb.save(excel_path)

def stress_test_constrained(w_co2_dict, w_air_dict, poly_dict, output_dir, min_order=1, max_order=4, current_date=None):
    if current_date is None:
        from datetime import datetime
        current_date = datetime.now().strftime('%Y-%m-%d')

    print("Running Constrained Stress Test...")

    X_values = np.linspace(0.50, 10.00, num=int((10.00 - 0.50) / 0.5) + 1)
    Y_values = np.linspace(0.05, 0.35, num=int((0.35 - 0.05) / 0.025) + 1)

    results_dict = {
        "Mass_Flow": [],
        "CO2_Concentration": [],
        "CO2_Th_PMF": [],
        "Air_Th_PMF": []
    }

    for degree in range(min_order, max_order+1):
        results_dict[f"CO2_Pred_Deg{degree}"] = []
        results_dict[f"Air_Pred_Deg{degree}"] = []

    for X in X_values:
        for Y in Y_values:
            results_dict["Mass_Flow"].append(X)
            results_dict["CO2_Concentration"].append(Y)
            results_dict["CO2_Th_PMF"].append(X * Y)
            results_dict["Air_Th_PMF"].append(X * (1 - Y))

            for degree in range(min_order, max_order+1):
                poly = poly_dict[degree]
                X_input = np.array([[X, Y]])
                Phi = poly.transform(X_input)
                co2_pred = float(Phi @ w_co2_dict[degree])
                air_pred = float(Phi @ w_air_dict[degree])
                results_dict[f"CO2_Pred_Deg{degree}"].append(co2_pred)
                results_dict[f"Air_Pred_Deg{degree}"].append(air_pred)

    df_st = pd.DataFrame(results_dict)

    for degree in range(min_order, max_order + 1):
        df_st[f"CO2_Dev_Deg{degree}"] = (df_st[f"CO2_Pred_Deg{degree}"] - df_st["CO2_Th_PMF"]) / df_st["CO2_Th_PMF"]
        df_st[f"Air_Dev_Deg{degree}"] = (df_st[f"Air_Pred_Deg{degree}"] - df_st["Air_Th_PMF"]) / df_st["Air_Th_PMF"]

    co2_pred_columns = [col for col in df_st.columns if "CO2_Pred_Deg" in col]
    air_pred_columns = [col for col in df_st.columns if "Air_Pred_Deg" in col]

    df_st["CO2_Best_guess"] = df_st[co2_pred_columns].apply(
        lambda row: row.dropna().loc[(row.dropna() - df_st.loc[row.name, "CO2_Th_PMF"]).abs().idxmin()]
        if not row.dropna().empty else np.nan,
        axis=1
    )
    df_st["Air_Best_guess"] = df_st[air_pred_columns].apply(
        lambda row: row.dropna().loc[(row.dropna() - df_st.loc[row.name, "Air_Th_PMF"]).abs().idxmin()]
        if not row.dropna().empty else np.nan,
        axis=1
    )

    ordered_columns = ["Mass_Flow", "CO2_Concentration", "CO2_Th_PMF", "CO2_Best_guess"] +                       co2_pred_columns + ["Air_Th_PMF", "Air_Best_guess"] + air_pred_columns
    df_st = df_st[ordered_columns]

    excel_path = os.path.join(output_dir, f'Data_Analysis_{current_date}.xlsx')
    mode = 'a' if os.path.exists(excel_path) else 'w'
    if mode == 'a':
        with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df_st.to_excel(writer, sheet_name='Stress_Test_Constrained', index=False)
    else:
        with pd.ExcelWriter(excel_path, engine='openpyxl', mode='w') as writer:
            df_st.to_excel(writer, sheet_name='Stress_Test_Constrained', index=False)

    wb = load_workbook(excel_path)
    ws = wb["Stress_Test_Constrained"]
    red_font = Font(color="FF0000")
    green_font = Font(color="008000")

    for col in range(3, len(df_st.columns) + 2):
        col_letter = ws.cell(row=1, column=col).column_letter
        ws.conditional_formatting.add(
            f"{col_letter}2:{col_letter}{len(df_st) + 1}",
            CellIsRule(operator="lessThan", formula=["0"], stopIfTrue=True, font=red_font),
        )

    for row_idx in range(2, len(df_st) + 2):
        co2_best_value = df_st.at[row_idx - 2, "CO2_Best_guess"]
        air_best_value = df_st.at[row_idx - 2, "Air_Best_guess"]

        for col_name in co2_pred_columns:
            if df_st.at[row_idx - 2, col_name] == co2_best_value:
                col_letter = ws.cell(row=1, column=df_st.columns.get_loc(col_name) + 1).column_letter
                ws.conditional_formatting.add(
                    f"{col_letter}{row_idx}:{col_letter}{row_idx}",
                    CellIsRule(operator="greaterThanOrEqual", formula=["0"], stopIfTrue=True, font=green_font),
                )
                break

        for col_name in air_pred_columns:
            if df_st.at[row_idx - 2, col_name] == air_best_value:
                col_letter = ws.cell(row=1, column=df_st.columns.get_loc(col_name) + 1).column_letter
                ws.conditional_formatting.add(
                    f"{col_letter}{row_idx}:{col_letter}{row_idx}",
                    CellIsRule(operator="greaterThanOrEqual", formula=["0"], stopIfTrue=True, font=green_font),
                )
                break

    wb.save(excel_path)

# --- 6. UTILITY FUNCTIONS ---
# format_polynomial_equation - Formats polynomial equations for output

def format_polynomial_equation(weights, powers, label):
    terms = []
    for coef, (i, j) in zip(weights, powers):
        if abs(coef) < 1e-8:
            continue
        term = f"({coef:.4f})"
        if i > 0:
            term += f"*X^{i}"
        if j > 0:
            term += f"*Y^{j}"
        terms.append(term)
    return f"{label} = " + " + ".join(terms)
